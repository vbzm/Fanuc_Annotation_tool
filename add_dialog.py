from PyQt5.QtWidgets import QHBoxLayout, QPushButton, QFormLayout, QLabel, QLineEdit, QVBoxLayout, QDialog, QMessageBox
from openpyxl.reader.excel import load_workbook


class AddDialog(QDialog):
    def __init__(self, parent=None):
        super(AddDialog, self).__init__(parent)
        self.init_ui(parent)

    def init_ui(self, parent):

        self.setWindowTitle('IP操作界面')

        hbox = QHBoxLayout()

        self.save_btn = QPushButton()
        self.save_btn.setText('保存')
        self.save_btn.clicked.connect(lambda: self.save_btn_click(parent))

        self.cancel_btn = QPushButton()
        self.cancel_btn.setText('取消')
        self.cancel_btn.clicked.connect(self.cancel_btn_click)

        hbox.addWidget(self.save_btn)
        hbox.addWidget(self.cancel_btn)

        '''表单布局'''
        fbox = QFormLayout()

        self.seq_lab = QLabel()
        self.seq_lab.setText('名称：')
        self.seq_text = QLineEdit()
        self.seq_text.setPlaceholderText('请输入设备名称')
        self.seq_text.setText(parent.text1)

        self.name_lab = QLabel()
        self.name_lab.setText('IP地址：')
        self.name_text = QLineEdit()
        self.name_text.setPlaceholderText('请输入IP地址')
        self.name_text.setText(parent.text2)

        self.age_lab = QLabel()
        self.age_lab.setText('HTTP账号：')
        self.age_text = QLineEdit()
        self.age_text.setPlaceholderText('请输入HTTP账号')
        self.age_text.setText(parent.text3)

        self.class_lab = QLabel()
        self.class_lab.setText('HTTP密码：')
        self.class_text = QLineEdit()
        self.class_text.setPlaceholderText('请输入HTTP密码')
        self.class_text.setText(parent.text4)

        self.socre_lab = QLabel()
        self.socre_lab.setText('备注：')
        self.socre_text = QLineEdit()
        self.socre_text.setPlaceholderText('请输入备注')
        self.socre_text.setText(parent.text5)

        fbox.addRow(self.seq_lab, self.seq_text)
        fbox.addRow(self.name_lab, self.name_text)
        fbox.addRow(self.age_lab, self.age_text)
        fbox.addRow(self.class_lab, self.class_text)
        fbox.addRow(self.socre_lab, self.socre_text)

        vbox = QVBoxLayout()
        vbox.addLayout(fbox)
        vbox.addLayout(hbox)

        self.setLayout(vbox)

    def save_btn_click(self, parent):
        if self.seq_text.text().strip() != '' and self.name_text.text().strip() != '':
            wb = load_workbook(parent.xlsx_path)
            ws = wb['IP']
            data = (self.seq_text.text(),
                    self.name_text.text(),
                    self.age_text.text(),
                    self.class_text.text(),
                    self.socre_text.text())
            if parent.change_ip_state:
                ws.cell(row=parent.row+2, column=1).value = data[0]
                ws.cell(row=parent.row+2, column=2).value = data[1]
                ws.cell(row=parent.row+2, column=3).value = data[2]
                ws.cell(row=parent.row+2, column=4).value = data[3]
                ws.cell(row=parent.row+2, column=5).value = data[4]
                parent.change_ip_state = False
            else:
                ws.append(data)
            wb.save(parent.xlsx_path)
            parent.get_ip_list()
            self.close()
        else:
            QMessageBox.information(self, "提示", "名称或IP地址不能为空！",
                                    QMessageBox.Yes)

    def cancel_btn_click(self):
        self.close()

    @staticmethod
    def get_add_dialog(parent=None):
        dialog = AddDialog(parent)
        return dialog.exec()
