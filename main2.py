# 原生的导入
import base64
import datetime
import os
import re
import sys
import threading
import time
import webbrowser
import socket

# 导入第三方库
from openpyxl import load_workbook, Workbook
import requests

# 导入基于PyQt5的Ui及一些工具
from PyQt5.QtCore import QThread, pyqtSignal, QCoreApplication, QStandardPaths
from PyQt5.QtGui import QTextCursor, QPalette, QColor, QIcon
from PyQt5.QtWidgets import QApplication, QWidget, QFileDialog, QMessageBox, QHeaderView, QAbstractItemView, \
    QTableWidgetItem, QDialog, QCheckBox, QLabel, QPushButton, QVBoxLayout, QDialogButtonBox

# 界面UI 以及多线程函数的导入
from add_dialog import AddDialog
from login import *
from mian_ui2 import *
from zb import *
from menu import *
from var_tool import *
from threads_func import IPColor

headers = {
    'Host': '192.168.0.99',
    'accept-encoding': 'gzip',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36',
    'Referer': 'http://192.168.0.99/karel/ComGet?sFc=33',
    'Accept-Encoding': 'gzip, deflate',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
    'Accept-Language': 'zh-CN,zh;q=0.9',
    'Authorization': 'Basic QUE6MTIz',
    # BASE64编码 'Authorization': 'Basic AA:123',
    'Connection': 'keep-aliv'
}

var_headers = {
    'Host': '127.0.0.1',
    'Referer': 'http://127.0.0.1/karel/ComSet?sComment=karel_e&sIndx=0&sFc=69',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.0.0 Safari/537.36',
}

rb_info = {
    'ip': '',
    'user': '',
    'version': '3.2',
    'debug_mode': False,
}

# 小铃铛闪烁功能实现 继承 QThread 并在其基础上实现的多线程
class WorkThread(QThread):
    timer = pyqtSignal()
    reset = pyqtSignal()
    set = pyqtSignal()

    def run(self):
        for i in range(0, 2):
            self.reset.emit()
            time.sleep(0.1)
            self.timer.emit()
            time.sleep(0.1)
        self.set.emit()

# 功能函数
class Func(QWidget):
    def __init__(self):
        super().__init__()
        # Ui定义
        self.debug = None
        self.setupUi = None
        self.window_y = None
        self.window_x = None
        self.MianUi = MianUI2()
        self.lineEdit_2 = None
        self.main_qt = None
        self.ui = None
        self.mian_help = None
        self.textEdit = None

        # 其他的变量定义
        self.robot_ip = rb_info['ip']
        self.path = None
        self.file = None
        self.connect = False
        self.GO_check = None
        self.num_io_check = None
        self.rb_io_check = None
        self.num_data_check = None
        self.pattern = None
        self.while_start_time = None
        self.del_start_time = None

        # openpyxl定义
        self.wb = Workbook()
        self.ws = None
        self.sheet = None
        self.di_list = []
        self.do_list = []
        self.r_list = []
        self.gi_list = []
        self.go_list = []
        self.ri_list = []
        self.ro_list = []
        self.checked_items = []
        self.all_note = {}

        # 多线程配置，用于小铃铛闪烁
        self.workThread = WorkThread()
        self.workThread.timer.connect(self.countTime)
        self.workThread.reset.connect(self.reset)
        self.workThread.set.connect(self.set_color)
        self.reset_color = "0, 255, 0"

    # 多线程函数配置
    def countTime(self):
        self.MianUi.label_23.setVisible(True)
        self.MianUi.label_3.setStyleSheet(f"background-color: rgb(255, 255, 255);")

    def reset(self):
        self.MianUi.label_23.setVisible(False)
        if self.reset_color:
            self.MianUi.label_3.setStyleSheet(f"background-color: rgb({self.reset_color});")

    def set_color(self):
        if self.reset_color == "255, 0, 0":
            self.reset_color = "0, 255, 0"

    # 公用函数
    def ping_ip(self, ip_str):
        try:
            result = requests.get(f"http://{ip_str}", timeout=2)
            if requests.get(f"http://{ip_str}/KAREL/COMMAIN", headers=headers).status_code == 401:
                self.reset_color = "255, 0, 0"
                self.new_info(f"连接失败， 请在{BASE_DIR}\\config\\config.py中配置正确用户名密码")
                return False
            self.new_info(f"连接成功，返回状态码：{result.status_code}\n再度点击通讯设置旁小按钮可断开连接")
            return True
        except requests.exceptions.ConnectionError:
            self.reset_color = "255, 0, 0"
            self.new_info("连接失败, 请检查IP地址是否正确")
            return False

    def new_info(self, text):
        self.workThread.start()

        now_info = self.MianUi.textEdit.toPlainText()
        new_info = f"{now_info}\n{text} -- {datetime.datetime.now().strftime('%H:%M:%S')}\n"
        # 将光标移动到最后
        self.MianUi.textEdit.setText(new_info)
        self.MianUi.textEdit.moveCursor(QTextCursor.End)

    def re_get_data(self, url_num):
        if url_num == 28:
            self.pattern = "<tr>\n.*\"center\">(.*?)</td>\n.*\n.*\n.*value=\"(.*?)\"\n.*\n.*\n.*\n.*value=\"(.*?)\""
        else:
            self.pattern = r"<tr>\n.*\">(.*?)<\/td>\n.*\n.*\n  value=\"(.*)\"\n.*\n.*\">(.*?)<\/td>\n.*\n.*\n  value=\"(.*)\"\n.*\n<\/tr>"
        html_con = requests.get(f"http://{self.robot_ip}/karel/ComGet?sFc={url_num}", headers=headers)
        html_con.encoding = "gb2312"
        return re.findall(self.pattern, html_con.text)

    def xlsx_while(self, sheet, data_list):
        self.new_info(f"正在写入{sheet}表")
        self.ws = self.wb.create_sheet(sheet, 0)
        if sheet == "数值寄存器":
            self.ws.append(["序号", "注释", "值"])
            self.ws.column_dimensions['A'].width = 15
            self.ws.column_dimensions['B'].width = 25
            self.ws.column_dimensions['C'].width = 25
            for data in data_list:
                self.ws.append([data[0], data[1], data[2]])
        else:
            self.ws.append(['Name', 'Data Type', 'PLC', 'RB1', 'RB1', '', 'Name', 'Data Type', 'PLC', 'RB1', 'RB1'])
            self.ws.column_dimensions['A'].width = 46
            self.ws.column_dimensions['B'].width = 10
            self.ws.column_dimensions['C'].width = 10
            self.ws.column_dimensions['D'].width = 10
            self.ws.column_dimensions['E'].width = 10
            self.ws.column_dimensions['G'].width = 46
            self.ws.column_dimensions['H'].width = 10
            self.ws.column_dimensions['I'].width = 10
            self.ws.column_dimensions['J'].width = 10
            self.ws.column_dimensions['K'].width = 10
            for data in data_list:
                self.ws.append(([data[3], '', '', data[2], '', '', data[1], '', '', data[0]]))
        self.new_info(f"{sheet}表写入完成")
        self.ws = None

    def xlsx_save(self):
        current_time = datetime.datetime.now().strftime('%Y-%m-%d(%H%M)')
        xlsx_name = f"{self.MianUi.lineEdit_3.text()}/RBdata_{current_time}.xlsx"
        # 删除 sheet
        self.wb.remove(self.wb['Sheet'])
        self.wb.save(xlsx_name)
        self.new_info(f"保存成功，文件路径：{xlsx_name}")
        self.wb = Workbook()  # 重新创建一个新的工作簿

    def TC(self, text):  # 关闭窗口触发以下事件
        a = QMessageBox.question(self, '提示',
                                 text,
                                 QMessageBox.Yes | QMessageBox.No,
                                 QMessageBox.No)
        if a == QMessageBox.Yes:
            return True
        else:
            return False

    def TC1(self, title, text):
        QMessageBox.information(self, f"{title}", f"{text}",
                                QMessageBox.Yes)

    """
    多线程槽函数
    """

    def del_success(self):
        self.new_info(f"注释删除成功, 耗时：{round((time.time() - self.del_start_time), 2)}秒")
        self.MianUi.pushButton_4.setEnabled(True)
        self.MianUi.pushButton_4.setStyleSheet("QPushButton{\n"
                                               "    background-color: rgb(41, 159, 255);\n"
                                               "color: rgb(255,255,255);\n"
                                               "border:0.5px solid rgb(0,0,0);\n"
                                               "border-radius:3px;\n"
                                               "}")

    def del_jdt(self, jdt_nums):
        print(jdt_nums)
        self.MianUi.progressBar_2.setValue(jdt_nums)

    def while_success(self):
        self.new_info(f"注释写入成功, 耗时：{round((time.time() - self.while_start_time), 2)}秒")
        self.MianUi.pushButton_2.setEnabled(True)
        self.MianUi.pushButton_2.setStyleSheet("QPushButton{\n"
                                               "    background-color: rgb(41, 159, 255);\n"
                                               "color: rgb(255,255,255);\n"
                                               "border:0.5px solid rgb(0,0,0);\n"
                                               "border-radius:3px;\n"
                                               "}")

    def while_jdt(self, jdt_nums):
        self.MianUi.progressBar.setValue(jdt_nums)

    """
    槽事件定义
    IP设置
    """

    # 注释写入按钮事件
    def while_note(self):
        self.robot_ip = rb_info['ip']
        if self.MianUi.lineEdit_2.text():
            wb = load_workbook(self.MianUi.lineEdit_2.text())
            print(wb.sheetnames)
            if "数字信号" in wb.sheetnames:
                self.sheet = wb["数字信号"]
                self.new_info("已找到数字信号表，正在读取数字信号表内容")
                self.do_list = self.find_note(0, 3)
                self.di_list = self.find_note(6, 9)
                self.new_info("共读取到DI注释{}条，DO注释{}条".format(len(self.di_list), len(self.do_list)))
            if "数值寄存器" in wb.sheetnames:
                self.sheet = wb["数值寄存器"]
                self.new_info("已找到数值寄存器表，正在读取数值寄存器表内容")
                self.r_list = self.find_note(1, 0)
                self.new_info("共读取到R数值寄存器注释{}条".format(len(self.r_list)))
            if "组信号" in wb.sheetnames:
                self.sheet = wb["组信号"]
                self.new_info("已找到组信号表，正在读取组信号表内容")
                self.gi_list = self.find_note(6, 9)
                self.go_list = self.find_note(0, 3)
                self.new_info("共读取到GI注释{}条，GO注释{}条".format(len(self.gi_list), len(self.go_list)))
            if "机器人信号" in wb.sheetnames:
                self.sheet = wb["机器人信号"]
                self.new_info("已找到机器人信号表，正在读取机器人信号表内容")
                self.ri_list = self.find_note(6, 9)
                self.ro_list = self.find_note(0, 3)
                self.new_info("共读取到RI注释{}条，RO注释{}条".format(len(self.ri_list), len(self.ro_list)))
            # 判断有没有任意一个表格是找到了数据的
            if self.do_list or self.di_list or self.r_list or self.gi_list or self.go_list or self.ri_list or self.ro_list:
                pass
            else:
                self.reset_color = "255, 0, 0"
                self.new_info("未找到任何一个表格是符合要求的, 请检查表格, 是否包含有\n1、数字信号\n2、数值寄存器\n3、组信号\n4、机器人信号\n\n此类命名的表格才可以被程序识别为注释表，你也可以选择导出机器人注释，观察正确的sheet命名方法。")
                return

            if rb_info['debug_mode']:
                # 以追加的方式向log.txt中写入di_list和do_list
                with open(BASE_DIR + "\\config\\log.txt", "a") as f:
                    f.write(f"{datetime.datetime.now().strftime('%Y-%m-%d(%H%M)')}di_list:{self.di_list}\n\n\n\n\n")
                    f.write(f"{datetime.datetime.now().strftime('%Y-%m-%d(%H%M)')}do_list:{self.do_list}\n\n\n\n\n")
                    f.write(f"{datetime.datetime.now().strftime('%Y-%m-%d(%H%M)')}r_list:{self.r_list}\n\n\n\n\n")
                    f.write(f"{datetime.datetime.now().strftime('%Y-%m-%d(%H%M)')}gi_list:{self.gi_list}\n\n\n\n\n")
                    f.write(f"{datetime.datetime.now().strftime('%Y-%m-%d(%H%M)')}go_list:{self.go_list}\n\n\n\n\n")
                    f.write(f"{datetime.datetime.now().strftime('%Y-%m-%d(%H%M)')}ri_list:{self.ri_list}\n\n\n\n\n")
                    f.write(f"{datetime.datetime.now().strftime('%Y-%m-%d(%H%M)')}ro_list:{self.ro_list}\n\n\n\n\n")

            note_name_list = []
            if self.di_list or self.do_list:
                note_name_list.append("数字信号")
            if self.r_list:
                note_name_list.append("数值寄存器")
            if self.gi_list or self.go_list:
                note_name_list.append("组信号")
            if self.ri_list or self.ro_list:
                note_name_list.append("机器人信号")

            if note_name_list:
                self.open_dialog(note_name_list, f"在IO表中找到了以下可写入的注释，请选择：")
                if self.checked_items:
                    self.new_info("开始写入注释")
                    self.while_start_time = time.time()
                    self.MianUi.pushButton_2.setEnabled(False)
                    self.MianUi.pushButton_2.setStyleSheet("QPushButton{\n"
                                                           "    background-color: rgb(255, 0, 0);\n"
                                                           "color: rgb(255,255,255);\n"
                                                           "border:0.5px solid rgb(0,0,0);\n"
                                                           "border-radius:3px;\n"
                                                           "}")
                    self.new_info("正在整合注释信息")

                    if "数字信号" in self.checked_items:
                        if self.do_list:
                            self.all_note.update({
                                9: self.do_list
                            })
                        if self.di_list:
                            self.all_note.update({
                                8: self.di_list
                            })
                    if "数值寄存器" in self.checked_items:
                        self.all_note.update({
                            1: self.r_list
                        })
                    if "组信号" in self.checked_items:
                        if self.go_list:
                            self.all_note.update({
                                11: self.go_list
                            })
                        if self.gi_list:
                            self.all_note.update({
                                10: self.gi_list
                            })
                    if "机器人信号" in self.checked_items:
                        if self.ri_list:
                            self.all_note.update({
                                6: self.ri_list
                            })
                        if self.ro_list:
                            self.all_note.update({
                                7: self.ro_list
                            })
                    self.checked_items = []

                    print(self.all_note)

                    from threads_func import Whilenote
                    while_threads = Whilenote(self.robot_ip, self.all_note, headers, self)
                    while_threads.while_success.connect(self.while_success)
                    while_threads.while_jdt.connect(self.while_jdt)
                    while_threads.start()

                    # 将用到的列表都复位
                    self.all_note = {}
                    self.do_list, self.di_list, self.r_list, self.go_list, self.gi_list, self.ri_list, self.ro_list = [], [], [], [], [], [], []

                else:
                    self.new_info("未选择任何注释信息")

            else:
                self.new_info('表格中没有注释信息')
        else:
            self.reset_color = "255, 0, 0"
            self.new_info("未选择文件")

    # 删除注释按钮事件
    def del_note(self):
        self.robot_ip = rb_info['ip']
        note_dict = {}
        for fc in [32, 33, 34]:
            url = f'http://{self.robot_ip}/karel/ComGet?sFc={fc}'
            html_con = requests.get(url, headers=headers)
            html_con.encoding = "gb2312"
            pattern = r"<tr>\n.*\">(.*?)<\/td>\n.*\n.*\n  value=\"(.*)\"\n.*\n.*\">(.*?)<\/td>\n.*\n.*\n  value=\"(.*)\"\n.*\n<\/tr>"
            note_result = re.findall(pattern, html_con.text)
            tem_list1 = []
            tem_list2 = []
            for note in note_result:  # 遍历正则匹配的列表
                if note[1] != '':
                    tem_list1.append(re.findall(r"\d+\.?\d*", note[0])[0])
                if note[3] != '':
                    tem_list2.append(re.findall(r"\d+\.?\d*", note[2])[0])
            if fc == 32:
                if tem_list1:
                    note_dict.update({
                        6: tem_list1
                    })
                if tem_list2:
                    note_dict.update({
                        7: tem_list2
                    })
            elif fc == 33:
                if tem_list1:
                    note_dict.update({
                        8: tem_list1
                    })
                if tem_list2:
                    note_dict.update({
                        9: tem_list2
                    })
            elif fc == 34:
                if tem_list1:
                    note_dict.update({
                        10: tem_list1
                    })
                if tem_list2:
                    note_dict.update({
                        11: tem_list2
                    })
        tem_list1 = []

        html_con = requests.get(f'http://{self.robot_ip}/karel/ComGet?sFc=28', headers=headers)
        html_con.encoding = "gb2312"
        pattern = r"<tr>\n.*?center\">(.*?)</td>\n.*?\n.*?\n.*?value=\"(.*?)\""
        note_result = re.findall(pattern, html_con.text)
        for note in note_result:  # 遍历正则匹配的列表
            if note[1] != '':
                tem_list1.append(re.findall(r"\d+\.?\d*", note[0])[0])
        if tem_list1:
            note_dict.update({
                1: tem_list1
        })
        # 把tem_list1删除
        del tem_list1, tem_list2
        print(note_dict)

        if note_dict:
            nums = 0
            note_name_list = []
            for note_d in note_dict:
                nums += len(note_dict[note_d])
                if note_d == 6:
                    note_name_list.append("机器人信号 RI")
                elif note_d == 7:
                    note_name_list.append("机器人信号 RO")
                elif note_d == 8:
                    note_name_list.append("数字信号 DI")
                elif note_d == 9:
                    note_name_list.append("数字信号 DO")
                elif note_d == 10:
                    note_name_list.append("组信号 GI")
                elif note_d == 11:
                    note_name_list.append("组信号 GO")
                elif note_d == 1:
                    note_name_list.append("数值寄存器")
            self.new_info(f"有{nums}条IO注释可以删除，请确认你的操作")
            self.open_dialog(note_name_list, f"在机器人中找到了以下可删除的注释，请选择：")
            if self.checked_items:
                self.del_start_time = time.time()
                print(self.checked_items)
                # 判断机器人信号 RI 不在列表中 并且note_dict存在6这个键
                if "机器人信号 RI" not in self.checked_items and 6 in note_dict:
                    note_dict.pop(6)
                if "机器人信号 RO" not in self.checked_items and 7 in note_dict:
                    note_dict.pop(7)
                if "数字信号 DI" not in self.checked_items and 8 in note_dict:
                    note_dict.pop(8)
                if "数字信号 DO" not in self.checked_items and 9 in note_dict:
                    note_dict.pop(9)
                if "组信号 GI" not in self.checked_items and 10 in note_dict:
                    note_dict.pop(10)
                if "组信号 GO" not in self.checked_items and 11 in note_dict:
                    note_dict.pop(11)
                if "数值寄存器" not in self.checked_items and 1 in note_dict:
                    note_dict.pop(1)

                print(note_dict)
                self.new_info(f"正在整合数据")
                self.MianUi.pushButton_4.setEnabled(False)
                self.MianUi.pushButton_4.setStyleSheet("QPushButton{\n"
                                                       "    background-color: rgb(255, 0, 0);\n"
                                                       "color: rgb(255,255,255);\n"
                                                       "border:0.5px solid rgb(0,0,0);\n"
                                                       "border-radius:3px;\n"
                                                       "}")
                from threads_func import Delnote
                del_threads = Delnote(self.robot_ip, note_dict, headers, self)
                del_threads.del_success.connect(self.del_success)
                del_threads.del_jdt.connect(self.del_jdt)
                del_threads.start()
            else:
                self.new_info("操作已取消")
        else:
            self.new_info("没有需要删除的注释,真干净啊！")

    # 读取注释按钮事件
    def read_note(self):
        self.robot_ip = rb_info['ip']
        self.GO_check = self.MianUi.checkBox.isChecked()
        self.num_io_check = self.MianUi.checkBox_2.isChecked()
        self.rb_io_check = self.MianUi.checkBox_3.isChecked()
        self.num_data_check = self.MianUi.checkBox_4.isChecked()

        if self.GO_check or self.num_io_check or self.rb_io_check or self.num_data_check:  # 判断是否勾选
            if self.MianUi.lineEdit_3.text():  # 判断是否输入了文件夹路径
                with open(BASE_DIR + "\\config\\log.txt", "a") as f:

                    if self.GO_check:
                        go_list = self.re_get_data(34)
                        if rb_info['debug_mode']:
                            f.write(f"{datetime.datetime.now().strftime('%Y-%m-%d(%H%M)')}go_list:{go_list}\n\n\n\n\n")
                        self.xlsx_while("组信号", go_list)

                    if self.num_io_check:
                        num_io_list = self.re_get_data(33)
                        if rb_info['debug_mode']:
                            f.write(
                                f"{datetime.datetime.now().strftime('%Y-%m-%d(%H%M)')}num_io_list:{num_io_list}\n\n\n\n\n")
                        self.xlsx_while("数字信号", num_io_list)

                    if self.rb_io_check:
                        rb_io_list = self.re_get_data(32)
                        if rb_info['debug_mode']:
                            f.write(
                                f"{datetime.datetime.now().strftime('%Y-%m-%d(%H%M)')}rb_io_list:{rb_io_list}\n\n\n\n\n")
                        self.xlsx_while("机器人信号", rb_io_list)

                    if self.num_data_check:
                        num_data_list = self.re_get_data(28)
                        if rb_info['debug_mode']:
                            f.write(
                                f"{datetime.datetime.now().strftime('%Y-%m-%d(%H%M)')}num_data_list:{num_data_list}\n\n\n\n\n")
                        self.xlsx_while("数值寄存器", num_data_list)

                self.xlsx_save()
            else:
                self.reset_color = "255, 0, 0"
                self.new_info("请输入文件储存路径")
        else:
            self.reset_color = "255, 0, 0"
            self.new_info("请至少选择一个选项")

    # 选择文件夹按钮事件
    def selectPath(self):
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly
        path_dialog = QFileDialog()
        desktop_path = QStandardPaths.writableLocation(QStandardPaths.DesktopLocation)
        path_dialog.setDirectory(desktop_path)
        self.path = path_dialog.getExistingDirectory(self, '请选择保存目录', options=options)
        if self.path:
            self.MianUi.lineEdit_3.setText(self.path)
            self.new_info(f'文件夹选择成功：{self.path}')

    # 选择文件按钮事件
    def selectFile(self):
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly
        file_dialog = QFileDialog()
        desktop_path = QStandardPaths.writableLocation(QStandardPaths.DesktopLocation)
        file_dialog.setDirectory(desktop_path)
        self.file, _ = file_dialog.getOpenFileName(self, '请选择标准模板IO表文件', '/path/to/default/',
                                                   "标准模板IO表文件(*.xlsx)", options=options)
        if self.file:
            self.MianUi.lineEdit_2.setText(self.file)
            self.new_info(f'文件已选择成功：{self.file}')

    # 全选按钮事件
    def all_select(self):
        self.MianUi.checkBox.setChecked(True)
        self.MianUi.checkBox_2.setChecked(True)
        self.MianUi.checkBox_3.setChecked(True)
        self.MianUi.checkBox_4.setChecked(True)

    def debug_mode(self):
        self.debug = self.MianUi.radioButton.isChecked()
        if self.debug:
            self.new_info("调试模式已开启, 请在config/log.txt查看日志")
            rb_info['debug_mode'] = True
        else:
            self.new_info("调试模式已关闭")
            rb_info['debug_mode'] = False

    def change_ip(self):
        if self.TC("请确认操作"):
            rb_info["ip"] = ""
            rb_info["user"] = ""
            self.close()
            main.show()
        else:
            self.new_info("操作已取消")

    # noinspection PyBroadException
    def new_up(self):
        try:
            r = requests.get(f"http://175.178.211.71:7890/api/version").json()
            if r['version'] == rb_info['version']:
                self.new_info("当前已是最新版本")
            else:
                self.new_info("发现新版本")
                self.new_info(f"版本号：{r['version']}")
                self.new_info(f"更新内容：{r['content']}")
                if self.TC("是否更新"):
                    self.new_info("已调用浏览器打开网页，请手动下载")
                    webbrowser.open(r['download_url'])
        except:
            self.TC1("提示", "当前无网络连接，无法检查更新")

    def del_path(self):
        self.MianUi.lineEdit_2.setText('')
        self.new_info("文件路径已清空")

    # 弹窗
    def open_dialog(self, SFB, info):
        # 创建一个弹窗
        dialog = QDialog(self)
        dialog.setWindowTitle("选择")
        # dialog.setFixedSize(500, 500)

        # 创建一个表单布局，用于放置选择控件和结果标签
        layout = QVBoxLayout()

        label1 = QLabel(info)
        layout.addWidget(label1)
        # 创建多个多选控件，并将它们添加到布局管理器中
        for i in SFB:
            layout.addWidget(QCheckBox(i))

        # 添加"确定"和"取消"按钮
        select_all = QPushButton("全选")
        select_all.setStyleSheet("background-color: #48c9b0; color: white; font-weight: bold;")
        ok_button = QPushButton("确定")
        ok_button.setStyleSheet("background-color: #48c9b0; color: white; font-weight: bold;")
        cancel_button = QPushButton("取消")
        cancel_button.setStyleSheet("background-color: #48c9b0; color: white; font-weight: bold;")

        # 创建一个水平布局，用于放置"确定"、"取消"按钮
        button_box = QDialogButtonBox(Qt.Horizontal, dialog)
        button_box.addButton(select_all, QDialogButtonBox.AcceptRole)
        button_box.addButton(ok_button, QDialogButtonBox.AcceptRole)
        button_box.addButton(cancel_button, QDialogButtonBox.RejectRole)

        # 创建一个垂直布局，用于放置表单布局和水平布局
        v_layout = QVBoxLayout()
        v_layout.addLayout(layout)
        v_layout.addWidget(button_box)

        dialog.setLayout(v_layout)

        def get_checked_items():
            # 遍历所有的多选控件，获取其isChecked状态，并将其存储到列表中
            self.checked_items = []
            for checked in range(layout.count()):
                widget = layout.itemAt(checked).widget()
                if isinstance(widget, QCheckBox) and widget.isChecked():
                    self.checked_items.append(widget.text())
            dialog.close()

        def set_checked():
            # 遍历layout中的多选控件，并将其设置为选中状态
            for checked in range(layout.count()):
                widget = layout.itemAt(checked).widget()
                if isinstance(widget, QCheckBox):
                    widget.setChecked(True)

        # 连接"确定"和"取消"按钮的单击事件
        ok_button.clicked.connect(get_checked_items)
        cancel_button.clicked.connect(dialog.close)
        select_all.clicked.connect(set_checked)

        # 显示弹窗
        dialog.exec_()

    def find_note(self, note_index, num_index):
        tem_list = []
        for row in self.sheet.iter_rows(min_row=2):
            if row[note_index].value:
                tem_list.append((row[note_index].value, re.findall(r"\d+", row[num_index].value)[0]))
        return tem_list

# 定义的鼠标类
class MouseMove(Func):
    def __init__(self):
        super().__init__()
        self.MianUi = None
        self.window_y = None
        self.window_x = None
        self.mouse_y = None
        self.mouse_x = None
        self.move_flag = None
        self.click_y = None
        self.click_x = None

    def mousePressEvent(self, evt):
        self.move_flag = True
        # 鼠标按下位置
        self.mouse_x = evt.globalX()
        self.mouse_y = evt.globalY()
        # 当前窗口位置
        self.window_x = self.x()
        self.window_y = self.y()

    def mouseReleaseEvent(self, evt):
        self.move_flag = False

    def mouseMoveEvent(self, evt):
        if self.move_flag:
            # 计算向量
            move_x = evt.globalX() - self.mouse_x
            move_y = evt.globalY() - self.mouse_y
            # 新的窗口位置
            new_x = self.window_x + move_x
            new_y = self.window_y + move_y
            # 移动窗口
            self.move(new_x, new_y)

    def enterEvent(self, QEvent):
        self.MianUi.label_24.setPixmap(QtGui.QPixmap(":/vx/img/icon.png"))
        self.MianUi.label_28.setText("微信公众号:不知名网友i,欢迎关注")

    def leaveEvent(self, QEvent):
        self.MianUi.label_24.setPixmap(QtGui.QPixmap(":/vx/img/公众号二维码.jpg"))
        self.MianUi.label_28.setText("铁汁来扫个码 走个关注呗~")

# 选择IP
class SelectIP(Func):
    def __init__(self):
        super().__init__()
        self.sey_bye()
        self.set_ip_color = None
        self.rb_info_ui = None
        self.xlsx_ip = []
        self.xlsx_path = BASE_DIR + "\\config\\ip_list.xlsx"
        self.MianUi = SelectMian()
        self.MianUi.setupUi(self)
        self.MianUi.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)  # 设置表格自适应宽度
        self.MianUi.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)  # 设置表格不可编辑
        self.MianUi.tableWidget.setSelectionBehavior(QAbstractItemView.SelectRows)  # 设置选中整行
        font = self.MianUi.tableWidget.horizontalHeader().font()  # 获取表头字体
        font.setBold(True)  # 设置字体加粗
        self.MianUi.tableWidget.horizontalHeader().setFont(font)  # 设置表头字体加粗
        self.MianUi.tableWidget.setPalette(QPalette(QColor(239, 239, 239)))  # 设置背景颜色
        self.MianUi.tableWidget.setAlternatingRowColors(True)  # 设置隔行变色
        self.MianUi.tableWidget.cellDoubleClicked.connect(self.use_ip)
        self.get_ip_list()

        self.text1 = None
        self.text2 = None
        self.text3 = None
        self.text4 = None
        self.text5 = None
        self.change_ip_state = False
        self.row = None

        self.http_user = None
        self.http_pass = None
        self.rb_ip = None
        self.rb_name = None
        
        self.ip_list = None

        self.desktop = QApplication.desktop()
        self.screenRect = self.desktop.screenGeometry()
        self.width = self.screenRect.width()

        self.ping_true_ip_list = []

    def sey_bye(self):
        # 判断sey_bye此文件是否存在
        if not os.path.exists(BASE_DIR + "\\config\\sey_bye.txt"):
            if self.TC('本程序由不知名网友i开发，微信公众号:不知名网友i,微信号：lh168b\n\n如果你获取此软件的渠道为收费渠道，请立即联系退款，如果你正在倒卖此软件，请停止你的行为！\n\n开发不易 感谢使用\n\n不要用于非法用途，不要用于商用。\n不要用于非法用途，不要用于商用。\n不要用于非法用途，不要用于商用。\n'):
                with open(BASE_DIR + "\\config\\sey_bye.txt", "w") as f:
                    f.write("1")
            else:
                sys.exit()


    # 获取IP列表 从xlsx
    def get_ip_list(self):
        wb = load_workbook(self.xlsx_path)
        self.sheet = wb['IP']
        for row in self.sheet.iter_rows(min_row=2):
            data = [
                row[0].value,
                row[1].value,
                row[2].value,
                row[3].value,
                row[4].value,
                ""
            ]
            self.xlsx_ip.append(data)
        self.query_data_list(self.xlsx_ip)
        self.xlsx_ip = []

    # 处理获取到的ip 启动多线程赋值表格中颜色
    def query_data_list(self, data):
        self.ip_list = []
        if len(data) != 0 and len(data[0]) != 0:
            self.MianUi.tableWidget.setRowCount(len(data))
            self.MianUi.tableWidget.setColumnCount(len(data[0]))
            for i in range(len(data)):
                self.ip_list.append((data[i][1], i))
                for j in range(len(data[0])):
                    self.MianUi.tableWidget.setItem(i, j, QTableWidgetItem(str(data[i][j])))
            # 设置窗口标题
            self.set_ip_color = IPColor(self.ip_list)
            self.set_ip_color.set_item_color.connect(self.set_item_color)
            self.set_ip_color.set_win_title.connect(self.set_win_title)
            self.set_ip_color.start()


    # 添加IP到IP列表
    def add_ip(self):
        AddDialog.get_add_dialog(self)

    # 删除IP
    def del_ip(self):
        row_select = self.MianUi.tableWidget.selectedItems()
        if len(row_select) != 0:
            row = row_select[0].row()
            if row == 0:
                self.TC1("提示", "默认IP不能删除")
            else:
                if self.TC("确认是否删除"):
                    wb = load_workbook(self.xlsx_path)
                    ws = wb['IP']
                    ws.delete_rows(row + 2)
                    wb.save(self.xlsx_path)
                    self.get_ip_list()
        else:
            self.TC1("提示", "请选择要删除的IP")

    # 更改IP
    def change_ip(self):
        # 获取光表所在行
        row_select = self.MianUi.tableWidget.selectedItems()
        if len(row_select) != 0:
            self.row = row_select[0].row()
            if self.row == 0:
                self.TC1("提示", "默认IP不能修改")
            else:
                self.text1 = self.MianUi.tableWidget.item(self.row, 0).text()
                self.text2 = self.MianUi.tableWidget.item(self.row, 1).text()
                self.text3 = self.MianUi.tableWidget.item(self.row, 2).text()
                self.text4 = self.MianUi.tableWidget.item(self.row, 3).text()
                self.text5 = self.MianUi.tableWidget.item(self.row, 4).text()
                self.change_ip_state = True
                AddDialog.get_add_dialog(self)
        else:
            self.TC1("提示", "请选择要修改的IP")

    # 使用IP
    def use_ip(self):
        # 获取光表所在行
        row_select = self.MianUi.tableWidget.selectedItems()
        if len(row_select) != 0:
            self.row = row_select[0].row()
            self.http_user = self.MianUi.tableWidget.item(self.row, 2).text()
            self.http_pass = self.MianUi.tableWidget.item(self.row, 3).text()
            if self.http_user == "None":
                self.http_user = "aa"
            if self.http_pass == "None":
                self.http_pass = "aa"
            headers["Authorization"] = "Basic " + base64.b64encode(
                (self.http_user + ":" + self.http_pass).encode("utf-8")).decode(
                "utf-8")
            self.rb_ip = self.MianUi.tableWidget.item(self.row, 1).text()
            if self.rb_ip == "local IP":
                import socket
                self.rb_ip = socket.gethostbyname(socket.gethostname())
            self.rb_name = self.MianUi.tableWidget.item(self.row, 0).text()
            if self.ping_ip(self.rb_ip) == 200:
                rb_info["ip"] = self.rb_ip
                rb_info["user"] = self.rb_name
                self.TC1("提示", "IP可用")
                self.close()
                menu_ui.move((self.width - menu_ui.width()), 0)
                menu_ui.setWindowFlags(Qt.WindowStaysOnTopHint)
                menu_ui.show()
            elif self.ping_ip(self.rb_ip) == 401:
                self.TC1("提示", "HTTP认证失败，检查HTTP用户名和密码")
            else:
                self.TC1("提示", "IP不可用")
        else:
            self.TC1("提示", "请选择正确的的IP")

    # 多线程回调函数 若IP可用则设置颜色
    def set_item_color(self, ip_color):
        # 设置颜色
        self.MianUi.tableWidget.item(ip_color[0], 5).setBackground(QColor(
            ip_color[1][0], ip_color[1][1], ip_color[1][2]
        ))

    # 多线程回调函数 设置窗口标题
    def set_win_title(self, txt):
        self.setWindowTitle(txt)

    # 公用函数 用于测试IP是否可用
    def ping_ip(self, ip_str):
        try:
            return requests.get(f"http://{ip_str}/KAREL/COMMAIN", headers=headers, timeout=2).status_code
        except requests.exceptions.ConnectionError:
            return False

    # 搜索ip按钮
    def find_ip(self):
        if self.TC("点击YES后软件将自动寻找网段内IP，过程可能会卡顿，请耐心等待，若长时间未响应，请强行终止并联系开发者。\n(同时只支持搜索一台设备，若电脑有打开仿真软件:RoboGuide，请关闭后再使用此功能。"):
            self.ping_true_ip_list = []
            self.find_all_ip()

            if self.ping_true_ip_list:
                if self.TC(f"已找到IP地址：{self.ping_true_ip_list[0][1]}\n设备名：{self.ping_true_ip_list[0][0]}\n点击确定即添加"):
                    self.text1 = self.ping_true_ip_list[0][0]
                    self.text2 = self.ping_true_ip_list[0][1]
                    self.text5 = self.ping_true_ip_list[0][2]
                    AddDialog.get_add_dialog(self)
                else:
                    self.TC1("提示", "操作已取消")
            else:
                self.TC1("提示", "IP未找到")

    # 用于def find_ip(self)函数
    def find_all_ip(self):
        myname = socket.getfqdn(socket.gethostname())
        myaddr = socket.gethostbyname(myname)
        args = "".join(myaddr)
        ip_pre = '.'.join(args.split('.')[:-1])

        threads = []
        for i in range(1, 256):
            ip = '%s.%s' % (ip_pre, i)
            threads.append(threading.Thread(target=self.test_ip, args={ip, }))
        for i in threads:
            i.start()
        for i in threads:
            i.join()

    # 用于def find_all_ip(self)函数 多线程
    def test_ip(self, ip_str):
        try:
            resp = requests.get(f"http://{ip_str}/KAREL/COMMAIN", headers=headers, timeout=2).status_code
            if resp == 200:
                web_rb_info = requests.get(f"http://{ip_str}/", headers=headers, timeout=2).text
                name = re.findall("Hostname: (.*?)<br>", web_rb_info)[0].strip()
                rb_no = re.findall("Robot No: (.*?)<br>", web_rb_info)[0].strip()
                self.ping_true_ip_list.append((name, ip_str, rb_no))
        except:
            return False

# 主窗口
class QtMian(MouseMove):
    def __init__(self):
        super().__init__()  # 调用父类的构造函数，创建窗口
        self.td_path = None
        self.MianUi = MianUI2()
        self.MianUi.setupUi(self)
        # self.setAttribute(Qt.WA_TranslucentBackground)  # 设置窗口背景透明
        self.new_info("已经准备好")
        self.setAcceptDrops(True)

    # 鼠标拖入事件
    def dragEnterEvent(self, evn):
        self.td_path = evn.mimeData().text().replace("file:///", "")
        # 鼠标放开函数事件
        evn.accept()

    def dropEvent(self, evn):
        # 判断文件后缀是不是xlsx
        if self.td_path.split('.')[-1] == 'xlsx':
            self.MianUi.lineEdit_2.setText(self.td_path)
            self.new_info(f'文件已选择成功：{self.td_path}')
        else:
            self.reset_color = "255, 0, 0"
            self.new_info("请拖入xlsx文件")

    def closeEvent(self, event):
        menu_ui.reset_main_Enabled()

# 机器人配置信息 —— 辅助窗口
class RbInfo(QWidget):
    def __init__(self):
        super().__init__()  # 调用父类的构造函数，创建窗口
        self.zb_mian = ZBMian()
        self.zb_mian.setupUi(self)
        self.label_num = 1
        self.label_name = None
        self.robot_ip = rb_info['ip']
        self.get_rb_info()
        self.red_txt = """setStyleSheet("background-color: rgb(245, 245, 245);"
                                        "border-top-right-radius:0px;"
                                        "border-top-left-radius:0px;"
                                        "border-bottom-right-radius:10px;"
                                        "border-bottom-left-radius:10px;"
                                        "color: rgb(255,39,125);")
"""
        self.blue_txt = """setStyleSheet("background-color: rgb(245, 245, 245);"
                                        "border-top-right-radius:0px;"
                                        "border-top-left-radius:0px;"
                                        "border-bottom-right-radius:10px;"
                                        "border-bottom-left-radius:10px;"
                                        "color: rgb(41,159,255);")
"""

        from threads_func import GetRbJOIN
        self.del_threads = GetRbJOIN(self.robot_ip, headers)
        self.del_threads.get_rb_join.connect(self.get_rb_join)
        self.del_threads.start()

    # 接收多线程函数返回的机器人关节数据
    def get_rb_join(self, rb_join, rb_world):
        rb_join = [i.strip() for i in rb_join[0]]
        for one_join in rb_join:
            self.label_name = f"self.zb_mian.label_{self.label_num * 2}"
            if float(one_join) <= 0:
                exec(f"{self.label_name}.{self.red_txt}")
            else:
                exec(f"{self.label_name}.{self.blue_txt}")
            exec(f"{self.label_name}.setText(one_join)")
            self.label_num = self.label_num + 1
        rb_world = [i.strip() for i in rb_world[0]]
        for one_world in rb_world:
            self.label_name = f"self.zb_mian.label_{self.label_num * 2}"
            if float(one_world) <= 0:
                exec(f"{self.label_name}.{self.red_txt}")
            else:
                exec(f"{self.label_name}.{self.blue_txt}")
            exec(f"{self.label_name}.setText(one_world)")
            self.label_num = self.label_num + 1
        self.label_num = 1

    # 获取机器人配置信息
    def get_rb_info(self):
        self.robot_ip = rb_info['ip']
        url = f'http://{self.robot_ip}/MD/SUMMARY.DG?_TEMPLATE=FRS:SUMMTMPC'
        r = requests.get(url, headers=headers).text
        config_info = re.findall(r'CONFIG::\n.*?\n([\w\W]*)MOTOR::', r)
        self.zb_mian.textEdit.setText(config_info[0].strip())

    # 浮点数字控件绑定事件
    def Spbox(self):
        self.del_threads.sleep_time = self.zb_mian.doubleSpinBox.value()

    def closeEvent(self, event):
        self.del_threads.quit_flag = True
        menu_ui.reset_rb_Enabled()

# 变量工具
class VarTool(Func):
    def __init__(self):
        super().__init__()
        self.data_list1 = None
        self.traversal_s = None
        self.init_s = None
        self.fu_s = None
        self.MianUi = VarToolUi()
        self.MianUi.setupUi(self)
        self.var_flag = False



    '''模糊搜索'''
    def return_fuzzy_search(self, search_req):
        self.MianUi.pushButton_4.setEnabled(True)
        self.MianUi.pushButton_4.setText("搜索")
        if search_req:
            self.MianUi.textEdit_2.setText(f"搜索到{len(search_req)}个变量")
            for i in search_req:
                self.MianUi.textEdit_2.setText(self.MianUi.textEdit_2.toPlainText() + f"\n路径：{i[0]} 详细：{i[1].strip()}")
        else:
            self.TC1("提示", "未找到匹配的变量")

    def fuzzy_search(self):
        se_str = self.MianUi.lineEdit_2.text()
        if se_str == "":
            self.TC1("提示", "请输入搜索内容")
            return
        else:
            self.robot_ip = rb_info['ip']
            self.MianUi.pushButton_4.setEnabled(False)
            self.MianUi.pushButton_4.setText("搜索中...")
            self.MianUi.textEdit_2.setText("")

            from threads_func import FuzzySearch
            self.fu_s = FuzzySearch(self.robot_ip, var_headers, se_str)
            self.fu_s.return_fuzzy_search.connect(self.return_fuzzy_search)
            self.fu_s.start()



    '''初始遍历'''
    def return_init_data(self):
        self.MianUi.pushButton.setEnabled(True)
        self.MianUi.pushButton.setText("记录")
        self.MianUi.label_4.setText("当前状态：初始程序成功。")
        self.var_flag = True

    def init_jd(self, jd):
        self.MianUi.label_4.setText(jd)

    def init_var_sql(self):
        self.robot_ip = rb_info['ip']
        self.MianUi.pushButton.setEnabled(False)
        self.MianUi.pushButton.setText("初始程序中...")
        self.MianUi.textEdit.setText("")
        self.MianUi.label_6.setText("当前状态：未开始")

        from threads_func import InitVarSql
        self.init_s = InitVarSql(self.robot_ip, BASE_DIR)
        self.init_s.return_init_data.connect(self.return_init_data)
        self.init_s.init_jd.connect(self.init_jd)
        self.init_s.start()


    '''遍历'''
    def return_traversal_data(self, data_list):
        self.data_list1 = data_list
        self.MianUi.pushButton_2.setEnabled(True)
        self.MianUi.pushButton_2.setText("遍历")
        self.MianUi.label_6.setText("当前状态：遍历成功，查看结果。")

        if self.data_list1:
            self.MianUi.textEdit.setText(f"搜索到{len(self.data_list1)}个变量")
            for i in self.data_list1:
                # 将i 去除头尾空格后换行追加到输入框
                self.MianUi.textEdit.setText(self.MianUi.textEdit.toPlainText() + f"\n{i.strip()}")
            self.MianUi.textEdit.setText(self.MianUi.textEdit.toPlainText() + f"\n遍历结束。")
        else:
            self.TC1("提示", "遍历失败，变量未更新。")

    def traversal_jd(self, jd):
        self.MianUi.label_6.setText(jd)

    def traversal_var(self):
        if not self.var_flag:
            self.TC1("提示", "请先执行初始程序。")
            return
        self.MianUi.pushButton_2.setEnabled(False)
        self.MianUi.pushButton_2.setText("遍历中...")
        self.robot_ip = rb_info['ip']

        from threads_func import TraversalVar
        self.traversal_s = TraversalVar(self.robot_ip, BASE_DIR)
        self.traversal_s.return_traversal_data.connect(self.return_traversal_data)
        self.traversal_s.traversal_jd.connect(self.traversal_jd)
        self.traversal_s.start()



    '''搜索'''
    def search_var(self):
        search_var_true = []
        if not self.data_list1 or self.data_list1 == []:
            self.TC1("提示", "请先执行遍历程序。")
            return
        se_str = self.MianUi.lineEdit.text()
        if se_str == "":
            self.TC1("提示", "请输入搜索内容")
            return
        else:
            for i in self.data_list1:
                if se_str in i:
                    # 将i 去除头尾空格后换行追加到search_var_true
                    search_var_true.append(i.strip())
            if search_var_true:
                # 将列表转为字符串
                search_var_true = "\n".join(search_var_true)
                self.TC1("搜索完成，请查看结果", f"{search_var_true}")
            else:
                self.TC1("提示", "未找到匹配的变量")


    '''重写关闭事件'''
    def closeEvent(self, event):
        menu_ui.reset_var_Enabled()

# 功能选择
class MenuUi(Func):
    def __init__(self):
        super().__init__()
        self.download_ls = None
        self.rb_info_ui = None
        self.var_tool_ui = None
        self.MianUi = MENU()
        self.MianUi.setupUi(self)
        self.MianUi.checkBox.setEnabled(False)
        self.MianUi.checkBox_2.setEnabled(False)
        self.MianUi.checkBox_3.setEnabled(False)
        self.ip = rb_info["ip"]

    # 机器人注释工具显示
    def mian_win_show(self):
        self.MianUi.checkBox.setEnabled(True)
        main_ui.MianUi.label_4.setText(f"当前连接IP地址：{rb_info['ip']}")
        main_ui.MianUi.label_5.setText(f"设备名称：{rb_info['user']}")
        main_ui.MianUi.label_6.setText(f"当前版本：{rb_info['version']}")
        main_ui.setWindowTitle(f"发那科注释工具 - {rb_info['version']}")
        main_ui.show()

    # 机器人配置信息显示
    # noinspection PyBroadException
    def rbinfo_win_show(self):
        self.MianUi.checkBox_2.setEnabled(True)
        try:
            self.rb_info_ui = RbInfo()
            self.rb_info_ui.setWindowIcon(QIcon(os.path.join(BASE_DIR, 'img', 'icon.ico')))
            self.rb_info_ui.show()
        except:
            self.TC1("提示", "机器人信息窗口打开失败")

    def fanuc_var_win_show(self):
        self.MianUi.checkBox_3.setEnabled(True)
        try:
            self.var_tool_ui = VarTool()
            self.var_tool_ui.setWindowIcon(QIcon(os.path.join(BASE_DIR, 'img', 'icon.ico')))
            self.var_tool_ui.show()
        except:
            self.TC1("提示", "发那科变量工具窗口打开失败")

    def download_all_ls(self):
        if self.TC("是否下载所有文件？\n点击是后确认LS文件保存目录"):
            self.ip = rb_info["ip"]
            ls_list_url = f"http://{self.ip}/MD/INDEX_TP.HTM"
            ls_html_response = requests.get(ls_list_url).text
            ls_list = re.findall(r'<TD align=center><A HREF=.*?HREF=".*?">(.*?)</A></TD>', ls_html_response)
            options = QFileDialog.Options()
            options |= QFileDialog.ReadOnly
            path_dialog = QFileDialog()
            desktop_path = QStandardPaths.writableLocation(QStandardPaths.DesktopLocation)
            path_dialog.setDirectory(desktop_path)
            self.path = path_dialog.getExistingDirectory(self, '请选择保存目录', options=options)
            if self.path:
                self.MianUi.pushButton_4.setEnabled(False)
                if self.TC(f"请确认文件是否保存到：{self.path}\nLS文件数量：{len(ls_list)}"):
                    from threads_func import DownloadLs
                    self.download_ls = DownloadLs(self.ip, headers, self.path, ls_list, self)
                    self.download_ls.download_success.connect(self.download_success)
                    self.download_ls.start()

    def download_success(self):
        self.MianUi.pushButton_4.setEnabled(True)
        self.TC1("提示", "下载完成")
        self.download_ls.quit()
        self.download_ls = None

    # 窗口置顶
    def mian_top(self):
        if self.MianUi.checkBox.isChecked():
            main_ui.setWindowFlags(Qt.WindowStaysOnTopHint)
            main_ui.show()
        else:
            main_ui.setWindowFlags(Qt.Window)
            main_ui.show()

    # 设置机器人信息窗口置顶
    def rbinfo_top(self):
        if self.MianUi.checkBox_2.isChecked():
            self.rb_info_ui.setWindowFlags(Qt.WindowStaysOnTopHint)
            self.rb_info_ui.show()
        else:
            self.rb_info_ui.setWindowFlags(Qt.Widget)
            self.rb_info_ui.show()

    # 设置变量工具窗口置顶
    def fanuc_var_top(self):
        if self.MianUi.checkBox_3.isChecked():
            self.var_tool_ui.setWindowFlags(Qt.WindowStaysOnTopHint)
            self.var_tool_ui.show()
        else:
            self.var_tool_ui.setWindowFlags(Qt.Widget)
            self.var_tool_ui.show()

    def reset_main_Enabled(self):
        self.MianUi.checkBox.setEnabled(False)

    def reset_rb_Enabled(self):
        self.MianUi.checkBox_2.setEnabled(False)

    def reset_var_Enabled(self):
        self.MianUi.checkBox_3.setEnabled(False)

    def close_all_win(self):
        # 判断checkBox是否可以被点击
        if self.MianUi.checkBox.isEnabled():
            main_ui.close()
        if self.MianUi.checkBox_2.isEnabled():
            self.rb_info_ui.close()
        if self.MianUi.checkBox_3.isEnabled():
            self.var_tool_ui.close()

    def closeEvent(self, event):
        exit(0)


if __name__ == '__main__':
    try:
        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        QCoreApplication.setAttribute(QtCore.Qt.AA_EnableHighDpiScaling)  # 高分辨率屏幕适配
        app = QApplication(sys.argv)
        main = SelectIP()
        main.setWindowIcon(QIcon(os.path.join(BASE_DIR, 'img', 'icon.ico')))
        main_ui = QtMian()
        main_ui.setWindowIcon(QIcon(os.path.join(BASE_DIR, 'img', 'icon.ico')))
        menu_ui = MenuUi()
        main.show()
        sys.exit(app.exec_())
    except Exception as e:
        print(e)
        app = QApplication(sys.argv)
        exit(0)
